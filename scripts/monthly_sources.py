"""30カテゴリ拡張 波2: 公的・業界団体の月次データ源の取得モジュール.

price_universe_check.py の series type からディスパッチされる（type "jmtba" → fetch_jmtba()）。
各 fetch_* は checker の parsed 規約
    {"value", "day_pct", "weekly_pct", "monthly_pct", "yoy_pct", "src_date", "layout"}
を返す（取れなければ None＝checker 側で parse_fail 記録）。src_date は "YYYY-MM"（公表月）で、
月次レーンの同一公表月dedupのキーになる。

取得実証・URL変化規則・更新ラグは 2026-08-02 SubAgent 調査が根拠（正本= docs/price-watch-universe.md §16m）:
- 日工会/SEAJ はPDF（ファイル名が毎月変わるため一覧ページから辿る）→ pypdf でテキスト化
- 自工会は素のGET不可・フォームPOST必須（約2ヶ月遅れの点に注意）
- 三鬼商事は robots.txt が xlsx/PDF を Disallow のため **HTMLページのみ** 使う（規約遵守）
- 農水省CSVは連番が不規則・非単調のため一覧の最初のリンク＝最新（中身の header 月で自己検証）
- JNTO xlsx は年別シート・「総数」行に値と前年同月比（伸率）が対で入る

パーサ（parse_*）は fetch から分離し、fixtureで selftest 可能にしている
（実行: python scripts/monthly_sources.py --selftest）。
"""
from __future__ import annotations

import csv
import io
import re
import sys
import time
from urllib.parse import urljoin

import requests

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")


def _get(url: str, **kw) -> requests.Response:
    resp = requests.get(url, headers={"User-Agent": UA}, timeout=60, **kw)
    resp.raise_for_status()
    return resp


def _pdf_text(data: bytes) -> str:
    """pypdf で全ページのテキストを連結する（依存は requirements.txt の pypdf）。"""
    from pypdf import PdfReader
    return "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(data)).pages)


def _num(s: str) -> float:
    return float(s.replace(",", ""))


# ---------------------------------------------------------------- 日工会（工作機械受注）

def parse_jmtba(txt: str) -> dict | None:
    """受注速報PDFのテキストから受注総額（百万円）・前月比・前年同月比を採る。

    表形式（2026-08-02 実測）: ヘッダ行「26/6月 前月比 前年同月比 2026年累計 前年同期比」の
    次行が受注総額の数値行「203,515 115.0 152.8 1,055,281 135.7」、続いて「うち内需」「うち外需」。
    自己検証: 内需+外需=総額（±0.5%）を要求（列取り違えの検知）。比は指数表記（115.0=+15%）。
    """
    if "受注総額" not in txt:
        return None
    d = re.search(r"(20\d{2})年(\d{1,2})月分", txt)
    if not d:
        return None
    # 表ヘッダ「… 前月比 前年同月比 …」以降にスコープする（PDF全体の先頭3数値行に
    # 依存すると、前段に同形式の行が増えたとき誤採取しうる・Codex S-6）
    hdr = re.search(r"前月比\s+前年同月比", txt)
    scope = txt[hdr.start():] if hdr else txt
    rows = re.findall(
        r"^(?:うち内需|うち外需)?\s*([0-9][0-9,]{3,})\s+([0-9]+\.[0-9])\s+([0-9]+\.[0-9])\s+"
        r"([0-9][0-9,]{3,})\s+([0-9]+\.[0-9])\s*$", scope, re.M)
    if len(rows) < 3:
        return None
    total, naiju, gaiju = (_num(r[0]) for r in rows[:3])
    if abs((naiju + gaiju) / total - 1) > 0.005:
        return None  # 列構成の変化＝静かに誤値を採らない
    return {"value": total, "day_pct": None, "weekly_pct": None,
            "monthly_pct": round(_num(rows[0][1]) - 100, 2),
            "yoy_pct": round(_num(rows[0][2]) - 100, 2),
            "src_date": f"{d.group(1)}-{int(d.group(2)):02d}", "layout": "jmtba_pdf_v1"}


def fetch_jmtba() -> dict | None:
    idx = _get("https://www.jmtba.or.jp/statistics/").text
    links = re.findall(r'href="([^"]*sokuhou\d{4}\.pdf)"', idx)
    if len(links) != 1:
        return None  # 「ちょうど1件」が崩れたらページ構造変化とみなす（実測2026-08-02）
    return parse_jmtba(_pdf_text(_get(urljoin("https://www.jmtba.or.jp/statistics/", links[0])).content))


# ---------------------------------------------------------------- SEAJ（半導体製造装置）

def parse_seaj(txt: str) -> dict | None:
    """SEAJ月度販売高速報PDFのテキストから販売高（百万円・3ヶ月移動平均）を採る。

    文面（2026-08-02 実測）: 「2026 年 6 月度の販売高は 513,610 百万円」
    「前月比 2.4％減…前年同月比 26.9%増」。％/% の表記揺れと増/減の符号を吸収する。
    """
    d = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月度の販売高は\s*([0-9,]+)\s*百万円", txt)
    if not d:
        return None

    def pct(label: str) -> float | None:
        m = re.search(label + r"[\s　]*[＋+]?\s*([0-9]+(?:\.[0-9]+)?)\s*[%％]\s*(増|減)", txt)
        if not m:
            return None
        return round(float(m.group(1)) * (1 if m.group(2) == "増" else -1), 2)

    yoy = pct("前年同月比")
    if yoy is None:
        # SEAJ の発火経路は前年同月比だけ（前月比は無効化済み）。表記変更で yoy が取れない
        # まま status=ok を返すと系列が**黙って永久沈黙**するため、fail-closed で parse_fail
        # に倒して要確認リストに出す（Codex S-7）
        return None
    return {"value": _num(d.group(3)), "day_pct": None, "weekly_pct": None,
            "monthly_pct": pct("前月比"), "yoy_pct": yoy,
            "src_date": f"{d.group(1)}-{int(d.group(2)):02d}", "layout": "seaj_pdf_v1"}


def fetch_seaj() -> dict | None:
    idx = _get("https://www.seaj.or.jp/statistics/").text
    m = re.search(r'href="(\d+\.pdf)"', idx)  # プレスリリース統計資料の先頭＝半導体速報（実測）
    if not m:
        return None
    return parse_seaj(_pdf_text(_get(urljoin("https://www.seaj.or.jp/statistics/", m.group(1))).content))


# ---------------------------------------------------------------- 自工会（四輪車生産台数）

_JAMA_URL = "https://jamaserv.jama.or.jp/newdb/prod4/prod4TsMkEntry.html?pass"


def parse_jama(html: str) -> dict | None:
    """自工会DBの応答HTMLから（YYYY年M月, 合計台数）の系列を採り、最新月とMoM/YoYを出す。

    フラット化すると「2026年1月|677,460|…|2026年5月|630,471|2026年6月|-」の対が並ぶ
    （2026-08-02 実測・未掲載月は「-」）。約2ヶ月遅れで掲載される点は系列noteに記載。
    """
    t = re.sub(r"<[^>]+>", "|", html)
    pairs = re.findall(r"(20\d{2})年(\d{1,2})月\|([0-9,]+|-)", re.sub(r"[\s|]+", "|", t))
    vals = {f"{y}-{int(m):02d}": _num(v) for y, m, v in pairs if v != "-"}
    if not vals:
        return None
    cur = max(vals)
    y, m = map(int, cur.split("-"))
    prev = f"{y}-{m - 1:02d}" if m > 1 else f"{y - 1}-12"
    prev_year = f"{y - 1}-{m:02d}"
    mom = (vals[cur] / vals[prev] - 1) * 100 if prev in vals else None
    yoy = (vals[cur] / vals[prev_year] - 1) * 100 if prev_year in vals else None
    return {"value": vals[cur], "day_pct": None, "weekly_pct": None,
            "monthly_pct": round(mom, 2) if mom is not None else None,
            "yoy_pct": round(yoy, 2) if yoy is not None else None,
            "src_date": cur, "layout": "jama_form_v1"}


def fetch_jama() -> dict | None:
    """JSESSIONID を先に取得（GETでcookie発行）→ フォーム値でPOST（fixture実測 2026-08-02）。

    フィールド名は `prod4TsMkEntryForm:doAction`（submit値=Server）等の JSF 形式で、
    key と value を取り違えるとブートストラップHTMLだけが返る（実害: 初回実装で再現）。
    """
    from datetime import datetime, timedelta
    now = datetime.now()
    entry = "https://jamaserv.jama.or.jp/newdb/prod4/prod4TsMkEntry.html"
    sess = requests.Session()
    sess.headers["User-Agent"] = UA
    sess.get(_JAMA_URL, timeout=60)  # JSESSIONID cookie の発行
    data = {
        "makerCd": "0", "chkSelCnd3": "0", "additionBase": "1", "additionInterval": "1",
        "termFrom": (now - timedelta(days=430)).strftime("%Y%m"),  # 14ヶ月＝前年同月を含める
        "termTo": now.strftime("%Y%m"),
        # requests は同名キーの多重送信を list で表現する（車種8区分の全選択）
        "car4Cd": ["100005", "100010", "100015", "200005",
                   "200010", "200015", "300005", "300010"],
        "prod4TsMkEntryForm:doAction": "Server",
        "prod4TsMkEntryForm/prod4/prod4TsMkEntry.html": "prod4TsMkEntryForm",
    }
    resp = sess.post(_JAMA_URL, data=data, timeout=60, headers={"Referer": entry})
    resp.raise_for_status()
    return parse_jama(resp.text)


# ---------------------------------------------------------------- 三鬼商事（東京オフィス）

def parse_miki(html: str) -> dict | None:
    """/rent/ のHTMLから都心5地区の平均賃料（円/坪）を採る。

    robots.txt が xlsx/PDF を Disallow しているため HTML のみ使う（2026-08-02 調査・規約遵守）。
    ブロック実測: 「都心5地区）…平均空室率 1.99 %…平均賃料 22,993 円…前月比 ▲148円」。
    前月比の記号は ▲/▼/△ をマイナスとみなす（和文IRの慣行）。空室率は参考として同 dict に載せる。
    """
    i = html.find("都心5地区")
    if i < 0:
        return None
    t = re.sub(r"<[^>]+>", " ", html[i:i + 4000])
    t = " ".join(t.split())
    rent = re.search(r"平均賃料\s*([0-9,]+)\s*円(?:.*?前月比\s*([▲▼△+＋]?)\s*([0-9,]+)\s*円)?", t)
    vac = re.search(r"平均空室率\s*([0-9.]+)\s*[%％]", t)
    d = re.search(r"(20\d{2})年(\d{1,2})月(?:末|時点|号|度)?", t) or \
        re.search(r"(20\d{2})年(\d{1,2})月", html)
    if not (rent and d):
        return None
    value = _num(rent.group(1))
    mom = None
    if rent.group(3):
        delta = _num(rent.group(3)) * (-1 if rent.group(2) in ("▲", "▼", "△") else 1)
        base = value - delta
        mom = round(delta / base * 100, 2) if base else None
    out = {"value": value, "day_pct": None, "weekly_pct": None, "monthly_pct": mom,
           "yoy_pct": None, "src_date": f"{d.group(1)}-{int(d.group(2)):02d}",
           "layout": "miki_html_v1"}
    if vac:
        out["vacancy_pct"] = float(vac.group(1))
    return out


def fetch_miki() -> dict | None:
    return parse_miki(_get("https://www.e-miki.com/rent/").text)


# ---------------------------------------------------------------- JA全農たまご（鶏卵）

def parse_tamago(html: str) -> dict | None:
    """月次相場ページ（東京M基準値・円/kg）の6年分×12ヶ月の表から最新月とMoM/YoYを採る。

    表構造（2026-08-02 実測）: ページに東京/大阪/名古屋/福岡の4表があるため id="tokyo" の
    ブロックにスコープする（東京M基準値が系列の定義）。年ヘッダは <th>2021<span>年</span></th>
    形式。各行が <th>N月</th> + 年数ぶんの <td>。未到来月は td が空。
    """
    i = html.find('id="tokyo"')
    if i < 0:
        # 東京アンカーが無いまま続行すると4都市の表を同じ年月キーへ流し込み、
        # HTML上で後にある都市の値が東京値を静かに上書きする（Codex C-2）。即 fail-closed
        return None
    end = html.find('id="osaka"', i)
    html = html[i:end if end > 0 else len(html)]
    years = [int(y) for y in re.findall(
        r"<th[^>]*>\s*(20\d{2})\s*(?:<span[^>]*>年</span>|年)?\s*</th>", html)]
    if not years:
        return None
    table: dict[str, float] = {}
    for m in re.finditer(r"<th[^>]*>\s*(\d{1,2})月\s*</th>((?:\s*<td[^>]*>[^<]*</td>)+)", html):
        month = int(m.group(1))
        cells = re.findall(r"<td[^>]*>\s*([0-9,]*)\s*</td>", m.group(2))
        for y, c in zip(years, cells):
            if c:
                table[f"{y}-{month:02d}"] = _num(c)
    if not table:
        return None
    cur = max(table)
    y, mth = map(int, cur.split("-"))
    prev = f"{y}-{mth - 1:02d}" if mth > 1 else f"{y - 1}-12"
    prev_year = f"{y - 1}-{mth:02d}"
    mom = (table[cur] / table[prev] - 1) * 100 if prev in table else None
    yoy = (table[cur] / table[prev_year] - 1) * 100 if prev_year in table else None
    return {"value": table[cur], "day_pct": None, "weekly_pct": None,
            "monthly_pct": round(mom, 2) if mom is not None else None,
            "yoy_pct": round(yoy, 2) if yoy is not None else None,
            "src_date": cur, "layout": "tamago_html_v1", "history": table}


def fetch_tamago() -> dict | None:
    parsed = parse_tamago(_get("https://www.jz-tamago.co.jp/business/souba/monthly/").text)
    if parsed:
        parsed.pop("history", None)  # 台帳行には最新値だけ載せる（historyは種まき用）
    return parsed


# ---------------------------------------------------------------- 農水省（米 相対取引価格）

_MAFF_LIST = "https://www.maff.go.jp/j/seisan/keikaku/soukatu/aitaikakaku.html"


def parse_rice(csv_text: str) -> dict | None:
    """相対取引価格CSVから全銘柄平均価格（円/60kg）とMoM/YoYを採る。

    ヘッダ（2026-08-02 実測）: 「8年5月_価格（7年産米）」等＝令和表記。値の行は
    「全銘柄平均価格、合計数量」。比は指数（99%=−1%）。
    列は**ヘッダ名で解決する**（固定位置 r[2]/r[8]/r[10] は列の挿入・並べ替えで別の数値を
    status=ok のまま採る静かな汚染経路になる・Codex C-1）。月列「<令和y>年<m>月_価格」を
    全部拾い、最大＝当月(value)・当月-1ヶ月＝前月・前年同月＝前年と解決し、比率は各価格列の
    **直後セルのヘッダ名**（対前月比/対前年比）を確認できた時だけ採る（確認できなければ
    None のまま＝発火しない）。
    """
    rows = list(csv.reader(io.StringIO(csv_text.lstrip("﻿"))))
    if not rows:
        return None
    hdr = rows[0]
    month_cols: dict[tuple[int, int], int] = {}
    for i, h in enumerate(hdr):
        m = re.match(r"(\d{1,2})年(\d{1,2})月_価格", h)
        if m:
            month_cols.setdefault((int(m.group(1)), int(m.group(2))), i)
    if not month_cols:
        return None
    cur = max(month_cols)  # (令和y, m) のタプル比較で最新月
    y, mth = cur
    prev_m = (y, mth - 1) if mth > 1 else (y - 1, 12)
    prev_y = (y - 1, mth)
    vrow = next((r for r in rows if r and r[0].startswith("全銘柄平均")), None)
    if not vrow:
        return None

    def cell(i: int | None) -> str:
        return vrow[i] if i is not None and i < len(vrow) else ""

    def pct_after(key: tuple[int, int], label: str) -> float | None:
        i = month_cols.get(key)
        if i is None or i + 1 >= len(hdr) or not hdr[i + 1].startswith(label):
            return None
        c = cell(i + 1)
        try:
            return round(_num(c.rstrip("%％")) - 100, 2) if c else None
        except ValueError:
            return None

    try:
        value = _num(cell(month_cols[cur]))
    except ValueError:
        return None
    return {"value": value, "day_pct": None, "weekly_pct": None,
            "monthly_pct": pct_after(prev_m, "対前月比"),
            "yoy_pct": pct_after(prev_y, "対前年比"),
            "src_date": f"{2018 + y}-{mth:02d}", "layout": "rice_csv_v1"}


def fetch_rice() -> dict | None:
    idx = _get(_MAFF_LIST).text
    m = re.search(r'href="([^"]*aitaikakaku-\d+\.csv)"', idx)  # 連番は不規則＝先頭リンクが最新（実測）
    if not m:
        return None
    resp = _get(urljoin(_MAFF_LIST, m.group(1)))
    resp.encoding = resp.apparent_encoding
    return parse_rice(resp.text)


# ---------------------------------------------------------------- JNTO（訪日外客数）

def parse_jnto_wb(wb) -> dict | None:
    """訪日外客数xlsx（年別シート・「総数」行に値と伸率=前年同月比が対）から最新月を採る。"""
    years = sorted((s for s in wb.sheetnames if re.fullmatch(r"20\d{2}", s)), reverse=True)
    for ysheet in years[:2]:  # 年初は最新年シートが空のことがあるため前年まで見る
        ws = wb[ysheet]
        header, total = None, None
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            if not row:
                continue
            # ヘッダ行は先頭セルが空（実測: [None,'1月','伸率','2月',…]）なので
            # 「N月 セルを含む行」で検出する
            if header is None and any(re.fullmatch(r"\d{1,2}月", str(c or "").strip())
                                      for c in row):
                header = row
            if str(row[0] or "").strip() == "総数":
                total = row
                break
        if not (header and total):
            continue
        months = []  # (month, value, yoy)
        for ci, h in enumerate(header):
            m = re.fullmatch(r"(\d{1,2})月", str(h or "").strip())
            if m and ci < len(total) and isinstance(total[ci], (int, float)):
                yoy = total[ci + 1] if ci + 1 < len(total) and \
                    isinstance(total[ci + 1], (int, float)) else None
                months.append((int(m.group(1)), float(total[ci]), yoy))
        if months:
            mth, value, yoy = months[-1]
            mom = (value / months[-2][1] - 1) * 100 if len(months) >= 2 else None
            return {"value": value, "day_pct": None, "weekly_pct": None,
                    "monthly_pct": round(mom, 2) if mom is not None else None,
                    "yoy_pct": round(yoy, 2) if yoy is not None else None,
                    "src_date": f"{ysheet}-{mth:02d}", "layout": "jnto_xlsx_v1"}
    return None


def fetch_jnto() -> dict | None:
    import openpyxl
    base = "https://www.jnto.go.jp/statistics/data/visitors-statistics/"
    idx = _get(base).text
    m = re.search(r'href="([^"]*/_files/\d{8}_1615-5\.xlsx)"', idx)  # 1615-5 は安定（実測）
    if not m:
        return None
    # 2026-08-17: 週次実行で "File is not a zip file" が1回発生（同URLを後刻叩くと正常＝一過性）。
    # xlsx は zip なので先頭が "PK"。違うものを openpyxl に渡すと原因の分からない例外になるため、
    # マジックバイトで判定し、1度だけ取り直す。それでも駄目なら**何が返ってきたか**を載せて失敗する
    url = urljoin(base, m.group(1))
    data = b""
    for attempt in (1, 2):
        data = _get(url).content
        if data[:2] == b"PK":
            break
        if attempt == 1:
            time.sleep(3)
    if data[:2] != b"PK":
        head = data[:80].decode("utf-8", errors="replace").replace("\n", " ")
        raise ValueError(f"JNTO xlsx が zip でない（{len(data)}バイト・先頭: {head!r}）")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    return parse_jnto_wb(wb)


# ---------------------------------------------------------------- e-Stat 生産動態（メモリ implied ASP）
# ASP分解レーン縮小継続（2026-08-04 ユーザー裁定・tasks/asp_decomposition_lane.md）。
# 探索経路は 2026-08-04 実測検証済み: 月ページ(tclass=12040605=月次確報)に statInfId が2つ
# （時系列表 h2daa<YYYYMM>_jikei.xlsx と機械統計生産能力指数表）。指数表でない方をDL。

_ESTAT_MONTH_PAGE = ("https://www.e-stat.go.jp/stat-search/files?page=1&layout=datalist&cycle=1"
                     "&toukei=00550200&tstat=000001022932&tclass1=000001058955&tclass2val=0"
                     "&year={y}0&month=12040605")
_ESTAT_DL = "https://www.e-stat.go.jp/stat-search/file-download?statInfId={sid}&fileKind=0"


def parse_estat_asp_memory(wb) -> dict | None:
    """生産動態・時系列表(実数)ワークブックから メモリ(調査票2360) の implied ASP を採る。

    ASP[円/個] = 販売金額[百万円] ÷ 販売数量[千個] × 1000。列は月ヘッダの正規表現で解決
    （12月号は13ヶ月窓・固定位置は不可＝backtest_v1で実測した罠）。fail-closed:
    実数表なし/メモリ行なし/金額数量ペア名不一致/数量0・非有限は None（欠測>誤記録）。
    """
    if "実数表" not in wb.sheetnames:
        return None
    it = wb["実数表"].iter_rows(values_only=True)
    next(it, None)
    header = next(it, None)
    if not header:
        return None
    cols = {str(v): i for i, v in enumerate(header) if v and re.fullmatch(r"20\d{4}", str(v))}
    if len(cols) < 7:  # QoQ(3/3)に6ヶ月+当月が要る
        return None
    amount = qty = None
    for r in it:
        if r[0] is None:
            continue
        if str(r[0]).strip() == "2360" and str(r[5]).strip() == "メモリ":
            item = str(r[6]).strip()
            if item == "販売金額":
                amount = {m: r[i] for m, i in cols.items()}
            elif item == "販売数量":
                qty = {m: r[i] for m, i in cols.items()}
    if not amount or not qty:
        return None
    months = sorted(cols)
    import math
    asp = {}
    for m in months:
        a, q = amount.get(m), qty.get(m)
        if (isinstance(a, (int, float)) and isinstance(q, (int, float))
                and math.isfinite(a) and math.isfinite(q) and q > 0):
            asp[m] = a / q * 1000.0
    ms = [m for m in months if m in asp]
    if len(ms) < 2:
        return None
    latest, prev = ms[-1], ms[-2]
    monthly_pct = round((asp[latest] / asp[prev] - 1) * 100, 2)
    qoq = None
    if len(ms) >= 6:
        recent = [asp[m] for m in ms[-3:]]
        earlier = [asp[m] for m in ms[-6:-3]]
        qoq = round((sum(recent) / 3 / (sum(earlier) / 3) - 1) * 100, 1)
    return {"value": round(asp[latest], 1), "day_pct": None, "weekly_pct": None,
            "monthly_pct": monthly_pct, "src_date": f"{latest[:4]}-{latest[4:]}",
            "layout": "estat_asp_memory_2360",
            "note_qoq3m": qoq, "n_months": len(ms)}


def fetch_estat_asp() -> dict | None:
    import datetime as _dt

    import openpyxl
    for year in (_dt.date.today().year, _dt.date.today().year - 1):  # 年初の未公表期は前年へ
        # fail-closed: 誤SID・非XLSX・壊れたWorkbook等の例外は握って次候補/None へ倒す
        # （誤った値を status=ok で記録するより欠測を選ぶ・Codex R1指摘）
        try:
            html = _get(_ESTAT_MONTH_PAGE.format(y=year)).text
            sids = re.findall(r"statInfId=(\d+)", html)
            if not sids:
                continue
            chosen = None
            for sid in dict.fromkeys(sids):
                seg_at = html.find(f"statInfId={sid}")
                if "指数表" not in html[max(0, seg_at - 2000):seg_at + 300]:
                    chosen = sid
                    break
            if chosen is None:
                continue
            wb = openpyxl.load_workbook(io.BytesIO(_get(_ESTAT_DL.format(sid=chosen)).content),
                                        data_only=True, read_only=True)
            try:
                parsed = parse_estat_asp_memory(wb)
            finally:
                wb.close()
            if parsed:
                return parsed
        except Exception:  # noqa: BLE001  (構造変化・DL失敗は欠測扱い＝発火しない)
            continue
    return None


# ---- 日銀の一括ダウンロード（CGPI/SPPI の全系列・品目別まで取れる唯一の無料経路） -------
# 2026-08-19 実証: 主要時系列ページ（mtshtml pr01_m_1/pr02_m_1）には**総平均など5系列しか無く**、
# 品目別（酪農品・菓子類・受託開発ソフトウェア等）は取れない。一方この一括ファイルには
# CGPI 3,042系列 / SPPI 513系列が入っており、データコードで1本を抜ける（実測）。
# §16v の宿題「日銀CGPI 細目の取得経路が無い」はこの経路で解消できる。
_BOJ_BULK_URL = {
    "cgpi": "https://www.stat-search.boj.or.jp/info/cgpi_m_jp.zip",   # 企業物価指数（月次）
    "sppi": "https://www.stat-search.boj.or.jp/info/sppi_m_jp.zip",   # 企業向けサービス価格指数（月次）
}
_BOJ_BULK_CACHE: dict[str, list[list[str]]] = {}   # 1回の実行で同じzipを何度も落とさない
# 指数の妥当レンジ（2020年=100）。桁違い・別系列の混入をここで弾く
_BOJ_INDEX_MIN, _BOJ_INDEX_MAX = 20.0, 1000.0
# 月次統計の公表ラグ（CGPIは翌月中旬・SPPIは翌々月下旬）。これを大きく超えたら鏡が凍っている
# CGPIは翌月中旬・SPPIは翌々月下旬の公表。ラグ2ヶ月＋余裕2ヶ月で 4 とする
# （6ヶ月だと、ファイル更新が止まっても半年間『正常だが鳴らない』状態になる・敵対レビュー指摘）
_BOJ_MAX_STALE_MONTHS = 4


def _boj_bulk_rows(dataset: str) -> list[list[str]]:
    """一括zipをCSV行として返す（同一実行内はキャッシュ）。"""
    if dataset in _BOJ_BULK_CACHE:
        return _BOJ_BULK_CACHE[dataset]
    import zipfile

    url = _BOJ_BULK_URL[dataset]
    zf = zipfile.ZipFile(io.BytesIO(_get(url).content))
    # 先頭ファイル決め打ちにしない（README等が同梱されたら別物を読む）。csv が1本の時だけ採用
    names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
    if len(names) != 1:
        raise ValueError(f"zip 内の csv が1本でない: {zf.namelist()}")
    raw = zf.read(names[0]).decode("shift_jis", errors="replace")
    rows = list(csv.reader(io.StringIO(raw)))
    _BOJ_BULK_CACHE[dataset] = rows
    return rows


def _ym_shift(ym: str, months: int) -> str:
    """"YYYYMM" を months だけずらす（前月・前年同月をキーで引くため）。"""
    y, m = int(ym[:4]), int(ym[4:])
    total = y * 12 + (m - 1) + months
    return f"{total // 12:04d}{total % 12 + 1:02d}"


def parse_boj_bulk(rows: list[list[str]], data_code: str,
                   today: str | None = None,
                   max_stale_months: int = _BOJ_MAX_STALE_MONTHS) -> dict | None:
    """一括CSVの行列から1系列を抜く。

    先頭行が `,,,202001,202002,...` の月ヘッダ、以降が
    `<データコード>,<統計名>,<系列名>,<値...>` の並び（2026-08-19 実測）。

    fail-closed の方針（この skill の落とし穴表どおり）:
      - データコードが無い / 値が1つも無い → None（空を成功にしない）
      - 月ヘッダが昇順・一意でない → None（列の意味が変わった可能性）
      - 指数レンジ（20〜1000）外 → None。**比較に使う前月・前年同月の値も同じ検査を通す**
        （最新値だけ見ていると、過去セルの異常値が巨大な前月比に化ける・2026-08-19 敵対レビュー）
      - 前月比・前年同月比は **月をキーで引く**。1ヶ月でも欠測があれば当該指標は None にする
        （「1つ前に存在する値」との比を『前月比』と呼ぶと、欠測時に別期間の騰落率になる）
      - 最新月が max_stale_months より古い → None（凍った鏡を新鮮な値として記録しない）

    Returns:
        {"value", "monthly_pct", "yoy_pct", "src_date"("YYYY-MM")} or None
    """
    if not rows:
        return None
    header = [h.strip() for h in rows[0]]
    months = [(i, h) for i, h in enumerate(header) if re.fullmatch(r"\d{6}", h)]
    if not months:
        return None
    labels = [h for _, h in months]
    if labels != sorted(labels) or len(set(labels)) != len(labels):
        return None   # 月が昇順・一意でない＝列の意味が変わった（推測しない）
    target = None
    for row in rows[1:]:
        if row and row[0].strip() == data_code:
            target = row
            break
    if target is None:
        return None   # コードが無い＝列を推測しない（黙って別系列を採らない）

    by_month: dict[str, float] = {}
    order: list[str] = []
    for i, ym in months:
        if i >= len(target):
            break
        cell = target[i].strip()
        if not cell:
            continue
        try:
            by_month[ym] = float(cell)
        except ValueError:
            continue
        order.append(ym)
    if not order:
        return None
    ym = order[-1]
    value = by_month[ym]

    def _sane(v: float | None) -> bool:
        return v is not None and _BOJ_INDEX_MIN <= v <= _BOJ_INDEX_MAX

    if not _sane(value):
        return None
    src_date = f"{ym[:4]}-{ym[4:]}"
    if today:
        age = (int(today[:4]) - int(ym[:4])) * 12 + (int(today[5:7]) - int(ym[4:]))
        if age > max_stale_months:
            return None
    prev = by_month.get(_ym_shift(ym, -1))
    year_ago = by_month.get(_ym_shift(ym, -12))
    return {
        "value": value,
        "monthly_pct": round((value / prev - 1) * 100, 2) if _sane(prev) and prev else None,
        "yoy_pct": round((value / year_ago - 1) * 100, 2) if _sane(year_ago) and year_ago else None,
        "src_date": src_date,
    }


def fetch_boj_bulk(dataset: str, data_code: str, today: str | None = None,
                   max_stale_months: int = _BOJ_MAX_STALE_MONTHS) -> dict | None:
    """日銀の一括ファイルから系列を1本取る（dataset: cgpi|sppi）。"""
    if dataset not in _BOJ_BULK_URL:
        return None
    try:
        return parse_boj_bulk(_boj_bulk_rows(dataset), data_code, today, max_stale_months)
    except Exception:  # noqa: BLE001  取得・解凍の失敗は欠測扱い（発火させない）
        return None



# ---------------------------------------------------------------- selftest

def _selftest() -> int:  # noqa: C901
    fails: list[str] = []

    def chk(name: str, cond: bool) -> None:
        print(("  ok " if cond else "  NG ") + name)
        if not cond:
            fails.append(name)

    # 日銀一括ファイル（2026-08-19 新設）。ネットワーク不要の固定行列で境界を検査する
    _boj_hdr = ["", "", "", "202604", "202605", "202606"]
    _boj_rows = [
        _boj_hdr,
        ["PRCS20_5201450001", "統計名", "品目/___受託開発ソフトウェア（除組込み）", "112.0", "113.0", "114.3"],
        ["PRCS20_ZERO", "統計名", "空っぽの系列", "", "", ""],
        ["PRCS20_HUGE", "統計名", "桁が違う系列", "1", "2", "999999"],
    ]
    b = parse_boj_bulk(_boj_rows, "PRCS20_5201450001", "2026-08")
    chk("boj_bulk 値/前月比/月",
        bool(b) and b["value"] == 114.3 and b["monthly_pct"] == 1.15 and b["src_date"] == "2026-06")
    chk("boj_bulk 前年比は12ヶ月ぶん無ければ None", bool(b) and b["yoy_pct"] is None)
    chk("boj_bulk 未知コードは None", parse_boj_bulk(_boj_rows, "PRCS20_NOPE", "2026-08") is None)
    chk("boj_bulk 値が1つも無ければ None（空を成功にしない）",
        parse_boj_bulk(_boj_rows, "PRCS20_ZERO", "2026-08") is None)
    chk("boj_bulk 指数レンジ外は None（桁違い・別系列の混入）",
        parse_boj_bulk(_boj_rows, "PRCS20_HUGE", "2026-08") is None)
    chk("boj_bulk 古すぎる最新月は None（凍った鏡を新鮮に見せない）",
        parse_boj_bulk(_boj_rows, "PRCS20_5201450001", "2027-06") is None)
    chk("boj_bulk 月ヘッダが無ければ None",
        parse_boj_bulk([["", "", ""], ["PRCS20_5201450001", "a", "b"]], "PRCS20_5201450001") is None)
    # 欠測月・異常な比較値・ヘッダ異常（2026-08-19 敵対レビュー NO-GO 1/2 の回帰）
    _gap = [["", "", "", "202603", "202604", "202605", "202606"],
            ["C", "統計名", "系列", "110.0", "", "113.0", "114.3"]]
    g = parse_boj_bulk(_gap, "C", "2026-08")
    chk("boj_bulk 欠測があっても前月比は隣接月のみで計算",
        bool(g) and g["monthly_pct"] == 1.15 and g["src_date"] == "2026-06")
    _gap2 = [["", "", "", "202603", "202604", "202605", "202606"],
             ["C", "統計名", "系列", "110.0", "111.0", "", "114.3"]]
    g2 = parse_boj_bulk(_gap2, "C", "2026-08")
    chk("boj_bulk 前月が欠測なら monthly_pct は None（2ヶ月前と比べない）",
        bool(g2) and g2["monthly_pct"] is None and g2["value"] == 114.3)
    _bad = [["", "", "", "202605", "202606"], ["C", "統計名", "系列", "0.5", "114.3"]]
    b2 = parse_boj_bulk(_bad, "C", "2026-08")
    chk("boj_bulk 比較する前月の値がレンジ外なら monthly_pct は None",
        bool(b2) and b2["monthly_pct"] is None)
    _unsorted = [["", "", "", "202606", "202605"], ["C", "統計名", "系列", "114.3", "113.0"]]
    chk("boj_bulk 月ヘッダが昇順でなければ None",
        parse_boj_bulk(_unsorted, "C", "2026-08") is None)
    chk("boj_bulk 既定の stale は4ヶ月（5ヶ月前は落とす）",
        parse_boj_bulk(_boj_rows, "PRCS20_5201450001", "2026-12") is None)

    u = parse_jmtba("受注総額\n2026年6月分　受注速報\n"
                    "26/6月 前月比 前年同月比 2026年累計 前年同期比\n"
                    "203,515 115.0 152.8 1,055,281 135.7\n"
                    "うち内需 58,016 128.0 145.5 272,892 122.8\n"
                    "うち外需 145,499 110.5 156.0 782,389 140.9\n")
    chk("jmtba 総額/比/月", bool(u) and u["value"] == 203515.0
        and u["monthly_pct"] == 15.0 and u["yoy_pct"] == 52.8 and u["src_date"] == "2026-06")
    u = parse_jmtba("受注総額\n2026年6月分\n26/6月\n100,000 115.0 152.8 1,055,281 135.7\n"
                    "うち内需 58,016 128.0 145.5 272,892 122.8\n"
                    "うち外需 145,499 110.5 156.0 782,389 140.9\n")
    chk("jmtba 内需+外需≠総額→None", u is None)

    s = parse_seaj("日本製 半導体製造装置 2026 年 6 月度の販売高は 513,610 百万円\n"
                   "前月比 2.4％減（…）、前年同月比 26.9%増（…）だった。")
    chk("seaj 値/符号（減=負・増=正）", bool(s) and s["value"] == 513610.0
        and s["monthly_pct"] == -2.4 and s["yoy_pct"] == 26.9 and s["src_date"] == "2026-06")

    j = parse_jama("<table><td>2025年5月</td><td>600,000</td><td>2026年4月</td><td>686,861</td>"
                   "<td>2026年5月</td><td>630,471</td><td>2026年6月</td><td>-</td></table>")
    chk("jama 最新月=未掲載スキップ・MoM/YoY", bool(j) and j["value"] == 630471.0
        and j["src_date"] == "2026-05" and abs(j["monthly_pct"] - (-8.21)) < 0.01
        and abs(j["yoy_pct"] - 5.08) < 0.01)

    mi = parse_miki('2026年6月末時点 <b>都心5地区）</b> <p>平均空室率</p><p>1.99</p>%'
                    '<p>平均賃料</p><p>22,993</p>円 <p>前月比</p> ▲ 148円')
    chk("miki 賃料/▲=マイナス/空室率", bool(mi) and mi["value"] == 22993.0
        and abs(mi["monthly_pct"] - (-0.64)) < 0.01 and mi.get("vacancy_pct") == 1.99)

    tokyo_tbl = ('<div id="tokyo"><tr><th>2025<span>年</span></th><th>2026<span>年</span></th></tr>'
                 "<tr><th>6月</th><td>240</td><td>290</td></tr>"
                 "<tr><th>7月</th><td>250</td><td>303</td></tr>"
                 "<tr><th>8月</th><td>260</td><td></td></tr></div>"
                 '<div id="osaka"><tr><th>2026<span>年</span></th></tr>'
                 "<tr><th>7月</th><td>999</td></tr></div>")
    ta = parse_tamago(tokyo_tbl)
    chk("tamago 最新月/MoM/YoY・大阪表を読まない", bool(ta) and ta["value"] == 303.0
        and ta["src_date"] == "2026-07"
        and abs(ta["monthly_pct"] - 4.48) < 0.01 and abs(ta["yoy_pct"] - 21.2) < 0.01)
    chk("tamago 東京アンカー無し→None（他都市値の上書き防止）",
        parse_tamago("<table><th>2026年</th><tr><th>7月</th><td>999</td></tr></table>") is None)

    rice_csv = ("産地,品種銘柄,8年5月_価格（7年産米）,8年5月_数量,年産平均,前年産平均,対前年比,"
                "8年4月_価格,対前月比,7年5月_価格,対前年比\n"
                '北海道,ななつぼし,"34,506","5,828","35,416","27,035",131%,"34,788",99%,"27,772",124%\n'
                '全銘柄平均価格、合計数量,,"33,164","53,817","35,812","25,179",142%,"33,447",99%,"27,649",120%\n')
    ri = parse_rice(rice_csv)
    chk("rice 全銘柄平均/令和→西暦/比→%", bool(ri) and ri["value"] == 33164.0
        and ri["src_date"] == "2026-05" and ri["monthly_pct"] == -1.0 and ri["yoy_pct"] == 20.0)
    # 列挿入への頑健性（ヘッダ名で解決・Codex C-1）: 先頭側に列を1本挿しても正しい値を採る
    rice_csv2 = ("産地,品種銘柄,備考,8年5月_価格（7年産米）,8年5月_数量,年産平均,前年産平均,対前年比,"
                 "8年4月_価格,対前月比,7年5月_価格,対前年比\n"
                 '全銘柄平均価格、合計数量,,メモ,"33,164","53,817","35,812","25,179",142%,'
                 '"33,447",99%,"27,649",120%\n')
    ri = parse_rice(rice_csv2)
    chk("rice 列挿入でも正しい列を解決", bool(ri) and ri["value"] == 33164.0
        and ri["monthly_pct"] == -1.0 and ri["yoy_pct"] == 20.0)
    # 比率列のヘッダ名が期待と違えば比を採らない（誤発火しない・値の記録は続く）
    ri = parse_rice(rice_csv.replace("対前月比", "対前月差"))
    chk("rice 比ヘッダ不一致→比None", bool(ri) and ri["value"] == 33164.0
        and ri["monthly_pct"] is None)
    # SEAJ: 唯一の発火経路（前年同月比）が取れない文面は fail-closed（Codex S-7）
    chk("seaj 前年同月比欠落→None",
        parse_seaj("2026 年 6 月度の販売高は 513,610 百万円 前月比 2.4％減 だった。") is None)

    print(f"[selftest] {'FAIL: ' + ', '.join(fails) if fails else 'all ok'}")
    return 1 if fails else 0


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        sys.exit(_selftest())
    # 手動疎通: python scripts/monthly_sources.py jmtba 等
    fn = globals().get(f"fetch_{sys.argv[1]}") if len(sys.argv) > 1 else None
    if fn:
        print(fn())
    else:
        print("usage: monthly_sources.py --selftest | <jmtba|seaj|jama|miki|tamago|rice|jnto>")
